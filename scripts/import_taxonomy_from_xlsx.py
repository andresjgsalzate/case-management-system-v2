"""Reconciliate CMS security_taxonomies with the client's operational
`Triage Eventos de Seguridad.xlsx` (sheet `Taxonomía v9`).

Behaviour:
  1. Parse the xlsx sheet (rows R7-R67 are the actual data; the rest is
     headers / formula scaffolding / repeated reference lists).
  2. For every distinct **parent**: ensure a root taxonomy exists
     (tenant_id NULL = global). New rows get auto-generated TUIC codes
     via the same slugify algorithm the frontend uses
     (UPPER-CASE-HYPHEN-ASCII). Existing rows matched by `name` keep
     their TUIC + description updated from the xlsx.
  3. For every **sub** row: ensure a child taxonomy exists, linked to
     its parent id, with the internal/external impacts mapped to the
     `ImpactLevel` slug enum (`bajo`/`medio`/`alto`/`critico`/
     `falso_positivo`).
  4. **Soft-delete** any active CMS taxonomy whose name is not in the
     xlsx (e.g. legacy DATA-EXFILTRATION, INSIDER-THREAT,
     ANOMALOUS-BEHAVIOR). Sets `is_active=false`; the row is preserved
     for audit + referential integrity (cases that point to it stay
     valid).

Idempotent: re-running produces no changes (lookups by name).

Usage:
  python3 scripts/import_taxonomy_from_xlsx.py
  python3 scripts/import_taxonomy_from_xlsx.py --dry-run  # just print

Requires `openpyxl` and `psycopg2-binary` (both auto-installed earlier
in the dev environment).
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata
import uuid
from contextlib import contextmanager
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import Json

# Where the xlsx lives. Assumes it's checked-in at repo root (or you can
# override with --xlsx <path>).
DEFAULT_XLSX_PATH = "Triage Eventos de Seguridad.xlsx"
SHEET_NAME = "Taxonomía v9"
DATA_START_ROW = 7   # R7 = first real taxonomy row ("Falso Positivo")
DATA_END_ROW = 67    # R67 = last real row; R68+ is meta + ArrayFormula scaffolding


# ─── Helpers ────────────────────────────────────────────────────────


def _load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    if os.path.exists("backend/.env"):
        with open("backend/.env", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                env[k] = v.strip('"').strip("'")
    return env


def _connect():
    url = _load_env().get("DATABASE_URL", "")
    if url.startswith("postgresql+asyncpg://"):
        url = url.replace("postgresql+asyncpg://", "postgresql://", 1)
    if not url:
        url = "postgresql://cms_user:cms_password@localhost:5433/cms_dev"
    return psycopg2.connect(url)


@contextmanager
def _tx(conn):
    try:
        yield conn.cursor()
        conn.commit()
    except Exception:
        conn.rollback()
        raise


def slugify_tuic(name: str) -> str:
    """Mirror of frontend `slugifyTuic` (TaxonomyEditModal.tsx). UPPER,
    ASCII-only, hyphen-separated, max 50 chars. Deterministic per name
    so re-runs of the import generate the same code for matching rows.
    """
    nfd = unicodedata.normalize("NFD", name)
    no_marks = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    upper = no_marks.upper()
    slug = re.sub(r"[^A-Z0-9]+", "-", upper)
    slug = slug.strip("-")
    return slug[:50]


def dedupe_tuic(base: str, taken: set[str]) -> str:
    """Append `-2`, `-3`, ... when `base` collides with an already-taken
    slug. Needed because the xlsx has name pairs that share slugs after
    diacritic stripping + case folding (e.g. parent "Código Malicioso"
    + sub "Código malicioso", or "Instalación o uso de software no
    autorizado" appearing as sub under both `Intrusiones` and
    `Violación de políticas`). Uniqueness is enforced by the partial
    index uq_taxonomy_tenant_tuic_active.
    """
    if base not in taken:
        taken.add(base)
        return base
    n = 2
    while True:
        # Make room for the suffix while still respecting the 50 char cap.
        suffix = f"-{n}"
        candidate = (base[: 50 - len(suffix)]).rstrip("-") + suffix
        if candidate not in taken:
            taken.add(candidate)
            return candidate
        n += 1


# Spanish impact label -> backend ImpactLevel slug (matches
# dtos.ImpactLevel Literal). Bajo/Medio/Alto/Crítico/Falso Positivo are
# the only values the xlsx uses in the impact columns.
IMPACT_MAP = {
    "bajo":            "bajo",
    "medio":           "medio",
    "alto":            "alto",
    "crítico":         "critico",
    "critico":         "critico",
    "falso positivo":  "falso_positivo",
}


def map_impact(value: str | None) -> str | None:
    if not value:
        return None
    return IMPACT_MAP.get(value.strip().lower())


# ─── Parse xlsx ──────────────────────────────────────────────────────


def parse_xlsx(xlsx_path: str) -> tuple[
    dict[str, str],                # parent_name -> description
    list[tuple[str, str, str | None, str | None]],  # (parent_name, sub_name, internal_impact, external_impact)
]:
    import warnings
    warnings.filterwarnings("ignore")

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    parents: dict[str, str] = {}      # name -> description (kept first non-empty)
    subs: list[tuple[str, str, str | None, str | None]] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx < DATA_START_ROW or row_idx > DATA_END_ROW:
            continue
        # Sheet layout: col A is empty (used for visual margin), real data
        # starts at col B. Hence row[1:6].
        cells = [str(c).strip() if c is not None else "" for c in row[1:6]]
        parent_name, description, sub_name, internal, external = (
            cells + ["", "", "", "", ""]
        )[:5]

        # Skip blank rows + meta markers (some rows under DATA_END could be junk).
        if not parent_name or any(
            m in parent_name for m in ("..inicio_", "..fin_", "<Seleccionar")
        ):
            continue
        if any(m in (sub_name or "") for m in ("ArrayFormula", "<Seleccione", "<seleccione")):
            continue

        # Capture the parent description on its first appearance.
        if parent_name not in parents:
            parents[parent_name] = description
        elif description and not parents[parent_name]:
            # Backfill if the first row happened to be missing description.
            parents[parent_name] = description

        # Subs row: always create if sub_name present.
        if sub_name:
            subs.append((
                parent_name, sub_name,
                map_impact(internal), map_impact(external),
            ))

    return parents, subs


# ─── DB sync ────────────────────────────────────────────────────────


def _pick_admin_user(cur) -> str:
    cur.execute("SELECT id FROM users ORDER BY created_at ASC LIMIT 1")
    row = cur.fetchone()
    if not row:
        sys.exit("No users in DB -- bootstrap a user first")
    return row[0]


def _upsert_root(
    cur,
    name: str,
    description: str,
    admin_user_id: str,
    *,
    dry_run: bool,
    taken_tuics: set[str],
) -> str:
    cur.execute(
        "SELECT id, tuic_code, description FROM security_taxonomies "
        "WHERE tenant_id IS NULL AND name = %s",
        (name,),
    )
    row = cur.fetchone()
    if row:
        rid, existing_tuic, existing_desc = row
        taken_tuics.add(existing_tuic)
        # Refresh description if changed (xlsx is source of truth).
        if (existing_desc or "") != (description or ""):
            print(f"  [UPDATE root] {name}: refreshing description")
            if not dry_run:
                cur.execute(
                    "UPDATE security_taxonomies "
                    "SET description = %s, is_active = TRUE, updated_at = NOW() "
                    "WHERE id = %s",
                    (description or None, rid),
                )
        else:
            # Still make sure it's active (re-activates anything we soft-deleted earlier).
            if not dry_run:
                cur.execute(
                    "UPDATE security_taxonomies SET is_active = TRUE WHERE id = %s",
                    (rid,),
                )
        return rid

    rid = str(uuid.uuid4())
    tuic = dedupe_tuic(slugify_tuic(name), taken_tuics)
    print(f"  [INSERT root] {name} (tuic={tuic})")
    if not dry_run:
        cur.execute(
            """
            INSERT INTO security_taxonomies
              (id, tenant_id, tuic_code, name, description, parent_id,
               attack_type, attack_subtype,
               internal_impact_context, external_impact_context,
               managed_by_team_id, default_case_type, requires_ticket,
               triage_mode, delegated_workflow_id, triage_timeout_seconds,
               tlp_default, prioritization_formula_id, mitre_techniques,
               is_active, forked_from_global_id, forked_from_global_at,
               created_at, updated_at, created_by, updated_by)
            VALUES
              (%s, NULL, %s, %s, %s, NULL,
               NULL, NULL,
               NULL, NULL,
               NULL, 'event', false,
               'auto', NULL, 300,
               'amber', NULL, %s,
               TRUE, NULL, NULL,
               NOW(), NOW(), %s, NULL)
            """,
            (rid, tuic, name, description or None, Json([]), admin_user_id),
        )
    return rid


def _upsert_sub(
    cur,
    name: str,
    parent_id: str,
    internal: str | None,
    external: str | None,
    admin_user_id: str,
    *,
    dry_run: bool,
    taken_tuics: set[str],
) -> str:
    cur.execute(
        "SELECT id, tuic_code, internal_impact_context, external_impact_context "
        "FROM security_taxonomies "
        "WHERE tenant_id IS NULL AND name = %s AND parent_id = %s",
        (name, parent_id),
    )
    row = cur.fetchone()
    if row:
        sid, existing_tuic, cur_int, cur_ext = row
        taken_tuics.add(existing_tuic)
        if cur_int != internal or cur_ext != external:
            print(
                f"    [UPDATE sub] {name}: "
                f"impacts {cur_int}/{cur_ext} -> {internal}/{external}"
            )
            if not dry_run:
                cur.execute(
                    "UPDATE security_taxonomies "
                    "SET internal_impact_context = %s, "
                    "    external_impact_context = %s, "
                    "    is_active = TRUE, updated_at = NOW() "
                    "WHERE id = %s",
                    (internal, external, sid),
                )
        else:
            if not dry_run:
                cur.execute(
                    "UPDATE security_taxonomies SET is_active = TRUE WHERE id = %s",
                    (sid,),
                )
        return sid

    sid = str(uuid.uuid4())
    tuic = dedupe_tuic(slugify_tuic(name), taken_tuics)
    print(f"    [INSERT sub] {name} (tuic={tuic}, parent={parent_id[:8]}…)")
    if not dry_run:
        cur.execute(
            """
            INSERT INTO security_taxonomies
              (id, tenant_id, tuic_code, name, description, parent_id,
               attack_type, attack_subtype,
               internal_impact_context, external_impact_context,
               managed_by_team_id, default_case_type, requires_ticket,
               triage_mode, delegated_workflow_id, triage_timeout_seconds,
               tlp_default, prioritization_formula_id, mitre_techniques,
               is_active, forked_from_global_id, forked_from_global_at,
               created_at, updated_at, created_by, updated_by)
            VALUES
              (%s, NULL, %s, %s, NULL, %s,
               NULL, NULL,
               %s, %s,
               NULL, 'event', false,
               'auto', NULL, 300,
               'amber', NULL, %s,
               TRUE, NULL, NULL,
               NOW(), NOW(), %s, NULL)
            """,
            (
                sid, tuic, name, parent_id,
                internal, external,
                Json([]), admin_user_id,
            ),
        )
    return sid


def _soft_delete_obsolete(
    cur, xlsx_names: set[str], *, dry_run: bool
) -> int:
    """Soft-delete active CMS global taxonomies whose name is not in the
    xlsx. Preserves the row (audit + FK from cases stays valid).
    """
    cur.execute(
        "SELECT id, name, tuic_code FROM security_taxonomies "
        "WHERE tenant_id IS NULL AND is_active = TRUE"
    )
    soft_deleted = 0
    for tid, name, tuic in cur.fetchall():
        if name not in xlsx_names:
            print(f"  [SOFT-DELETE] {name} (tuic={tuic})")
            if not dry_run:
                cur.execute(
                    "UPDATE security_taxonomies "
                    "SET is_active = FALSE, updated_at = NOW() "
                    "WHERE id = %s",
                    (tid,),
                )
            soft_deleted += 1
    return soft_deleted


# ─── Entrypoint ─────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=DEFAULT_XLSX_PATH)
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Print what would change without committing.",
    )
    args = parser.parse_args()

    if not Path(args.xlsx).exists():
        sys.exit(f"xlsx not found at: {args.xlsx}")

    parents, subs = parse_xlsx(args.xlsx)
    print(f"Parsed xlsx: {len(parents)} parents, {len(subs)} subs")
    all_names = set(parents.keys()) | {s[1] for s in subs}

    conn = _connect()
    try:
        with _tx(conn) as cur:
            admin_id = _pick_admin_user(cur)
            print(f"Admin user_id: {admin_id}\n")

            # Seed taken_tuics with EVERY active CMS tuic up front so new
            # rows don't collide even with legacy taxonomies we're not
            # touching this run.
            cur.execute(
                "SELECT tuic_code FROM security_taxonomies "
                "WHERE tenant_id IS NULL AND is_active = TRUE"
            )
            taken_tuics: set[str] = {r[0] for r in cur.fetchall()}

            print("== ROOTS ==")
            parent_ids: dict[str, str] = {}
            for pname, pdesc in parents.items():
                parent_ids[pname] = _upsert_root(
                    cur, pname, pdesc, admin_id,
                    dry_run=args.dry_run, taken_tuics=taken_tuics,
                )

            print("\n== SUBS ==")
            for parent_name, sub_name, internal, external in subs:
                pid = parent_ids.get(parent_name)
                if not pid:
                    print(f"    [WARN] sub '{sub_name}' has no resolved parent")
                    continue
                _upsert_sub(
                    cur, sub_name, pid, internal, external, admin_id,
                    dry_run=args.dry_run, taken_tuics=taken_tuics,
                )

            print("\n== SOFT-DELETE OBSOLETE ==")
            removed = _soft_delete_obsolete(cur, all_names, dry_run=args.dry_run)

            print("\n== SUMMARY ==")
            cur.execute(
                "SELECT COUNT(*) FILTER (WHERE is_active), COUNT(*) "
                "FROM security_taxonomies WHERE tenant_id IS NULL"
            )
            active, total = cur.fetchone()
            print(f"  Global taxonomies: {active} active / {total} total")
            print(f"  Soft-deleted this run: {removed}")
            if args.dry_run:
                print("\n  (dry-run -- nothing committed)")
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
